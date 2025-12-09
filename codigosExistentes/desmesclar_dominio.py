"""
Utilitario para gerar uma copia desmesclada dos relatorios do Domonio.

Uso:
    python desmesclar_dominio.py "CAMINHO\\DOMINIO...xls[x]" [saida.xlsx]

Comportamento:
- Se a origem for .xls, converte para .xlsx via pandas (sem formatacao) e salva.
- Se a origem for .xlsx, remove merges com openpyxl e preenche as celulas do range
  com o valor da celula superior esquerda.

Resultado: uma planilha sem merges, facilitando leitura por pandas.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def desmesclar_ws(ws):
    """Desfaz merges e replica o valor da celula superior esquerda em todas as celulas do range."""
    merges = list(ws.merged_cells.ranges)
    for m in merges:
        ws.unmerge_cells(str(m))
        tl_val = ws.cell(m.min_row, m.min_col).value
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                cell = ws.cell(r, c)
                if cell.value in (None, ""):
                    cell.value = tl_val


def process_xlsx(src: Path, dst: Path):
    wb = load_workbook(src)
    for ws in wb.worksheets:
        desmesclar_ws(ws)
    wb.save(dst)


def process_xls(src: Path, dst: Path):
    # Leitura sem merges; apenas converte para xlsx
    df = pd.read_excel(src, sheet_name=0, header=None)
    df.to_excel(dst, index=False, header=False)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    src = Path(sys.argv[1])
    if not src.exists():
        print(f"Arquivo nao encontrado: {src}")
        sys.exit(1)

    if len(sys.argv) >= 3:
        dst = Path(sys.argv[2])
    else:
        dst = src.with_name(src.stem + "_limpo.xlsx")

    ext = src.suffix.lower()
    try:
        if ext == ".xlsx":
            process_xlsx(src, dst)
        elif ext == ".xls":
            process_xls(src, dst)
        else:
            print(f"Extensao nao suportada: {ext}")
            sys.exit(1)
    except Exception as exc:
        print(f"Erro ao processar: {exc}")
        sys.exit(1)

    print(f"Arquivo gerado: {dst}")


if __name__ == "__main__":
    main()
